import os
import sys
import json
import subprocess
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def clean_string(s):
    """
    Removes surrogate characters (range D800-DFFF) that cause XML encoding failures in openpyxl.
    """
    if not isinstance(s, str):
        return s
    return ''.join(c for c in s if not (0xD800 <= ord(c) <= 0xDFFF))

def export_to_excel(analysis_data, excel_path, boundary_layer=None):
    """
    Exports flat land use planning table to Excel.
    """
    layers = {clean_string(k): v for k, v in analysis_data.get('layers', {}).items()}
    boundary_area = analysis_data.get('boundary_area', 1.0)
    
    if boundary_layer:
        boundary_layer = clean_string(boundary_layer)
        
    # Calculate road area dynamically: Boundary Area - Sum(All other layers excluding Road and Boundary)
    other_sum = sum(v for k, v in layers.items() if k != '도로' and k != boundary_layer)
    road_area = max(0.0, boundary_area - other_sum)
    layers['도로'] = road_area
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "토지이용계획표"
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_name = "맑은 고딕"
    title_font = Font(name=font_name, size=16, bold=True, color="1F4E79")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=11)
    bold_font = Font(name=font_name, size=11, bold=True)
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    sum_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "토지이용계획 면적산출표"
    ws["A1"].font = title_font
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 40
    
    # Headers
    headers = ["구분 (레이어명)", "면적 (㎡)", "구성비 (%)", "비고"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[3].height = 25
    
    # ─── Sum Row Header (placeholder: formulas filled after data rows are written) ───
    ws["A4"] = "합계"
    ws["A4"].font = bold_font
    ws["A4"].fill = sum_fill
    ws["A4"].alignment = align_center
    ws["A4"].border = thin_border
    
    ws["D4"] = ""
    ws["D4"].font = bold_font
    ws["D4"].fill = sum_fill
    ws["D4"].alignment = align_center
    ws["D4"].border = thin_border
    ws.row_dimensions[4].height = 25
    
    # Write Layer Data rows (start at row 5)
    DATA_START_ROW = 5
    current_row = DATA_START_ROW
    for name, area_val in sorted(layers.items()):
        # Skip boundary layer itself from row data if it's listed (already in Sum)
        if boundary_layer and name == boundary_layer:
            continue
        if area_val <= 0:
            continue
            
        c1 = ws.cell(row=current_row, column=1, value=name)
        c2 = ws.cell(row=current_row, column=2, value=round(area_val))
        c3 = ws.cell(row=current_row, column=3, value=f"=B{current_row}/B4")
        c4 = ws.cell(row=current_row, column=4, value="")
        
        c1.font = data_font
        c1.alignment = align_left
        c1.border = thin_border
        
        c2.font = data_font
        c2.alignment = align_right
        c2.number_format = '#,##0'
        c2.border = thin_border
        
        c3.font = data_font
        c3.alignment = align_right
        c3.number_format = '0.0%'
        c3.border = thin_border
        
        c4.font = data_font
        c4.border = thin_border
        
        ws.row_dimensions[current_row].height = 20
        current_row += 1
    
    # ─── Now fill in SUM formulas into B4 and C4 after all rows are written ───
    last_data_row = current_row - 1  # last row that has data
    if last_data_row >= DATA_START_ROW:
        sum_range = f"B{DATA_START_ROW}:B{last_data_row}"
        ws["B4"] = f"=SUM({sum_range})"
        ws["C4"] = f"=SUM(C{DATA_START_ROW}:C{last_data_row})"
    else:
        ws["B4"] = 0
        ws["C4"] = 0
    
    ws["B4"].font = bold_font
    ws["B4"].fill = sum_fill
    ws["B4"].alignment = align_right
    ws["B4"].number_format = '#,##0'
    ws["B4"].border = thin_border
    
    ws["C4"].font = bold_font
    ws["C4"].fill = sum_fill
    ws["C4"].alignment = align_right
    ws["C4"].number_format = '0.0%'
    ws["C4"].border = thin_border
        
    # Auto adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    wb.save(excel_path)

def export_to_hwp(analysis_data, hwp_path, boundary_layer=None):
    """
    Exports flat land use planning table to HWP.
    Launches a worker subprocess with timeout to avoid blocking/freezing the GUI.
    """
    # Clean input strings
    cleaned_layers = {clean_string(k): v for k, v in analysis_data.get('layers', {}).items()}
    cleaned_boundary_layer = clean_string(boundary_layer) if boundary_layer else None
    boundary_area = analysis_data.get('boundary_area', 1.0)
    
    # Calculate road area dynamically: Boundary Area - Sum(All other layers excluding Road and Boundary)
    other_sum = sum(v for k, v in cleaned_layers.items() if k != '도로' and k != cleaned_boundary_layer)
    road_area = max(0.0, boundary_area - other_sum)
    cleaned_layers['도로'] = road_area
    
    hwp_data = {
        'layers': cleaned_layers,
        'boundary_area': boundary_area,
        'boundary_layer': cleaned_boundary_layer
    }
    
    # Save parameters to a temp JSON file
    temp_json = hwp_path + ".temp.json"
    with open(temp_json, 'w', encoding='utf-8') as f:
        json.dump(hwp_data, f, ensure_ascii=False, indent=2)
        
    try:
        worker_script = os.path.join(os.path.dirname(__file__), 'create_hwp.py')
        
        result = subprocess.run(
            [sys.executable, worker_script, temp_json, hwp_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=8
        )
        
        if result.returncode != 0:
            raise RuntimeError(f"HWP Worker Error: {result.stderr}")
            
    except subprocess.TimeoutExpired:
        raise TimeoutError("한글(HWP) 연동 대기 시간을 초과했습니다. 한글 프로그램 보안 팝업을 허용했는지 확인해 주세요.")
    finally:
        # Clean up temp file
        if os.path.exists(temp_json):
            try:
                os.remove(temp_json)
            except Exception:
                pass
